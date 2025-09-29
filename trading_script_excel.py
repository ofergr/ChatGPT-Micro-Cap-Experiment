#!/usr/bin/env python3
"""
Excel-compatible wrapper for the trading script.
This script uses Excel files (.xlsx) for the portfolio data instead of CSV.
"""

import sys
import os
from pathlib import Path
import pandas as pd
import logging

# Add the current directory to the path to import the original trading script
sys.path.insert(0, str(Path(__file__).parent))

# Import from the original trading script
from trading_script import (
    daily_results, load_benchmarks, last_trading_date, set_data_dir, _log_initial_state,
    _ensure_df, trading_day_window, download_price_data, log_sell, log_manual_buy, log_manual_sell
)
import numpy as np
from excel_portfolio_utils import read_portfolio_excel, write_portfolio_excel

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format=' %(asctime)s - %(filename)s:%(lineno)d - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def load_latest_portfolio_state_excel(data_dir: Path) -> tuple[pd.DataFrame, float]:
    """Load the latest portfolio state from Excel file."""
    
    excel_path = data_dir / "chatgpt_portfolio_update.xlsx"
    
    if not excel_path.exists():
        # Try to find CSV file and convert it
        csv_path = data_dir / "chatgpt_portfolio_update.csv"
        if csv_path.exists():
            print(f"📊 Converting CSV to Excel: {csv_path} -> {excel_path}")
            df = pd.read_csv(csv_path)
            write_portfolio_excel(df, excel_path)
        else:
            raise FileNotFoundError(
                f"Could not find portfolio file at {excel_path} or {csv_path}.\\n"
                "Make sure you have portfolio data available."
            )
    
    logger.info(f"Reading Excel file: {excel_path}")
    df = read_portfolio_excel(excel_path)
    
    if df.empty:
        logger.warning("Portfolio Excel file is empty")
        return pd.DataFrame(columns=["ticker", "shares", "stop_loss", "buy_price", "cost_basis"]), 0.0
    
    # Get the most recent date's portfolio state
    # Handle mixed date formats by converting everything consistently
    df['Date_Clean'] = pd.to_datetime(df['Date'], errors='coerce')
    
    # Also try to parse string dates manually
    for idx, row in df.iterrows():
        if pd.isna(df.loc[idx, 'Date_Clean']) and pd.notna(row['Date']):
            try:
                # Try parsing as string
                df.loc[idx, 'Date_Clean'] = pd.to_datetime(str(row['Date']))
            except:
                pass
    
    # Remove rows with invalid dates or empty ticker
    df = df.dropna(subset=['Date_Clean'])
    df = df[df['Ticker'].notna() & (df['Ticker'] != '')]
    
    if df.empty:
        return pd.DataFrame(columns=["ticker", "shares", "stop_loss", "buy_price", "cost_basis"]), 0.0
    
    # Get today's date in string format to match what we expect
    today_str = last_trading_date().strftime("%Y-%m-%d")
    today_dt = pd.to_datetime(today_str)
    
    # Look for today's data first
    today_entries = df[df['Date_Clean'] == today_dt]
    
    if not today_entries.empty:
        print(f"📅 Found {len(today_entries)} entries for today ({today_str})")
        latest_entries = today_entries.copy()
    else:
        # Fall back to most recent date
        latest_date = df['Date_Clean'].max()
        latest_entries = df[df['Date_Clean'] == latest_date].copy()
        print(f"📅 Using most recent date: {latest_date.strftime('%Y-%m-%d')} ({len(latest_entries)} entries)")
    
    # Extract individual stock positions (exclude TOTAL)
    stocks = latest_entries[latest_entries['Ticker'] != 'TOTAL'].copy()
    
    # Get cash balance from TOTAL row - use pandas method (handles formulas better)
    cash = 0.0
    total_row = latest_entries[latest_entries['Ticker'] == 'TOTAL']
    if not total_row.empty and pd.notna(total_row['Cash Balance'].iloc[0]):
        cash = float(total_row['Cash Balance'].iloc[0])
        print(f"💰 Found cash balance ${cash:.2f} from latest TOTAL row")
    
    # Convert to the format expected by the trading script
    # For each ticker, get the LATEST row (in case of multiple rows like SELL then BUY)
    portfolio_data = []
    
    # Group by ticker and get the last (most recent) row for each ticker
    if not stocks.empty:
        # Sort by index to ensure we get the latest entry for each ticker
        stocks_sorted = stocks.sort_index()
        latest_per_ticker = stocks_sorted.groupby('Ticker').tail(1)
        
        for _, row in latest_per_ticker.iterrows():
            if pd.notna(row['Ticker']) and pd.notna(row['Shares']) and float(row['Shares']) > 0:
                portfolio_data.append({
                    'ticker': row['Ticker'],
                    'shares': float(row['Shares']),
                    'buy_price': float(row['Buy Price']) if pd.notna(row['Buy Price']) else 0.0,
                    'cost_basis': float(row['Cost Basis']) if pd.notna(row['Cost Basis']) else 0.0,
                    'stop_loss': float(row['Stop Loss']) if pd.notna(row['Stop Loss']) else 0.0,
                })
    
    portfolio_df = pd.DataFrame(portfolio_data) if portfolio_data else pd.DataFrame(columns=["ticker", "shares", "stop_loss", "buy_price", "cost_basis"])
    
    logger.info(f"Successfully loaded portfolio with {len(portfolio_df)} positions and ${cash} cash")
    return portfolio_df, cash

def process_portfolio_excel(
    portfolio: pd.DataFrame,
    cash: float,
    data_dir: Path,
    interactive: bool = True,
) -> tuple[pd.DataFrame, float]:
    """Excel-compatible version of process_portfolio that saves results to Excel."""
    
    today_date = last_trading_date().date()  # Use date object for consistency
    portfolio_df = _ensure_df(portfolio)

    results: list[dict] = []
    total_value = 0.0
    total_pnl = 0.0
    
    # Track manual sells for cash balance updates
    manual_sells: dict[str, float] = {}  # ticker -> cash from sale

    # ------- Interactive trade entry -------
    if interactive:
        while True:
            print(portfolio_df)
            action = input(
                f""" You have {cash} in cash.
Would you like to log a manual trade? Enter 'b' for buy, 's' for sell, or press Enter to continue: """
            ).strip().lower()

            if action == "b":
                ticker = input("Enter ticker symbol: ").strip().upper()
                order_type = input("Order type? 'm' = market-on-open, 'l' = limit: ").strip().lower()

                try:
                    shares = float(input("Enter number of shares: "))
                    if shares <= 0:
                        raise ValueError
                except ValueError:
                    print("Invalid share amount. Buy cancelled.")
                    continue

                if order_type == "m":
                    # Handle stop loss percentage with empty input support
                    stop_loss_input = input("Enter stop loss percentage (e.g., 8 for 8%, or press Enter to skip): ").strip()
                    try:
                        if stop_loss_input == "":
                            stop_loss_pct = 0.0
                        else:
                            stop_loss_pct = float(stop_loss_input)
                            if stop_loss_pct < 0:
                                raise ValueError
                    except ValueError:
                        print("Invalid stop loss percentage. Buy cancelled.")
                        continue

                    s, e = trading_day_window()
                    fetch = download_price_data(ticker, start=s, end=e, auto_adjust=False, progress=False)
                    data = fetch.df
                    if data.empty:
                        print(f"MOO buy for {ticker} failed: no market data available (source={fetch.source}).")
                        continue

                    o = float(data["Open"].iloc[-1]) if "Open" in data else float(data["Close"].iloc[-1])
                    exec_price = round(o, 2)
                    # Calculate stop loss price from percentage
                    stop_loss = round(exec_price * (1 - stop_loss_pct / 100), 2) if stop_loss_pct > 0 else 0.0
                    
                    notional = exec_price * shares
                    if notional > cash:
                        print(f"MOO buy for {ticker} failed: cost {notional:.2f} exceeds cash {cash:.2f}.")
                        continue
                    
                    if stop_loss_pct > 0:
                        print(f"Stop loss set at {stop_loss_pct}% = ${stop_loss:.2f}")

                    # Log to trade log CSV
                    log = {
                        "Date": today_date.strftime("%Y-%m-%d"),
                        "Ticker": ticker,
                        "Shares Bought": shares,
                        "Buy Price": exec_price,
                        "Cost Basis": notional,
                        "PnL": 0.0,
                        "Reason": "MANUAL BUY MOO - Filled",
                    }
                    # Add to trade log
                    import os
                    trade_log_path = data_dir / "chatgpt_trade_log.csv"
                    if os.path.exists(trade_log_path):
                        logger.info("Reading CSV file: %s", trade_log_path)
                        df_log = pd.read_csv(trade_log_path)
                        logger.info("Successfully read CSV file: %s", trade_log_path)
                        if df_log.empty:
                            df_log = pd.DataFrame([log])
                        else:
                            df_log = pd.concat([df_log, pd.DataFrame([log])], ignore_index=True)
                    else:
                        df_log = pd.DataFrame([log])
                    logger.info("Writing CSV file: %s", trade_log_path)
                    df_log.to_csv(trade_log_path, index=False)
                    logger.info("Successfully wrote CSV file: %s", trade_log_path)

                    # Update portfolio
                    rows = portfolio_df.loc[portfolio_df["ticker"].astype(str).str.upper() == ticker.upper()]
                    if rows.empty:
                        new_trade = {
                            "ticker": ticker,
                            "shares": float(shares),
                            "stop_loss": float(stop_loss),
                            "buy_price": float(exec_price),
                            "cost_basis": float(notional),
                        }
                        if portfolio_df.empty:
                            portfolio_df = pd.DataFrame([new_trade])
                        else:
                            portfolio_df = pd.concat([portfolio_df, pd.DataFrame([new_trade])], ignore_index=True)
                    else:
                        idx = rows.index[0]
                        cur_shares = float(portfolio_df.at[idx, "shares"])
                        cur_cost = float(portfolio_df.at[idx, "cost_basis"])
                        new_shares = cur_shares + float(shares)
                        new_cost = cur_cost + float(notional)
                        avg_price = new_cost / new_shares if new_shares else 0.0
                        portfolio_df.at[idx, "shares"] = new_shares
                        portfolio_df.at[idx, "cost_basis"] = new_cost
                        portfolio_df.at[idx, "buy_price"] = avg_price
                        portfolio_df.at[idx, "stop_loss"] = float(stop_loss)

                    cash -= notional
                    print(f"Manual BUY MOO for {ticker} filled at ${exec_price:.2f} ({fetch.source}).")
                    continue

                elif order_type == "l":
                    try:
                        buy_price = float(input("Enter buy LIMIT price: "))
                        if buy_price <= 0:
                            raise ValueError
                    except ValueError:
                        print("Invalid buy price. Limit buy cancelled.")
                        continue
                    
                    # Handle stop loss percentage with empty input support
                    stop_loss_input = input("Enter stop loss percentage (e.g., 8 for 8%, or press Enter to skip): ").strip()
                    try:
                        if stop_loss_input == "":
                            stop_loss_pct = 0.0
                        else:
                            stop_loss_pct = float(stop_loss_input)
                            if stop_loss_pct < 0:
                                raise ValueError
                        
                        # Calculate stop loss price from percentage
                        stop_loss = round(buy_price * (1 - stop_loss_pct / 100), 2) if stop_loss_pct > 0 else 0.0
                        if stop_loss_pct > 0:
                            print(f"Stop loss set at {stop_loss_pct}% = ${stop_loss:.2f}")
                        else:
                            print("No stop loss set.")
                    except ValueError:
                        print("Invalid stop loss percentage. Limit buy cancelled.")
                        continue

                    cash, portfolio_df = log_manual_buy(
                        buy_price, shares, ticker, stop_loss, cash, portfolio_df
                    )
                    continue
                else:
                    print("Unknown order type. Use 'm' for market-on-open or 'l' for limit.")
                    continue

            if action == "s":
                try:
                    ticker = input("Enter ticker symbol: ").strip().upper()
                    shares = float(input("Enter number of shares to sell (LIMIT): "))
                    sell_price = float(input("Enter sell LIMIT price: "))
                    if shares <= 0 or sell_price <= 0:
                        raise ValueError
                except ValueError:
                    print("Invalid input. Manual sell cancelled.")
                    continue

                original_cash = cash
                cash, portfolio_df = log_manual_sell(
                    sell_price, shares, ticker, cash, portfolio_df
                )
                # Track cash from this manual sell
                cash_from_sell = cash - original_cash
                if cash_from_sell > 0:
                    manual_sells[ticker] = manual_sells.get(ticker, 0) + cash_from_sell
                continue

            break  # proceed to pricing

    # ------- Daily pricing + stop-loss execution -------
    # Only process stocks that are currently in the portfolio (shares > 0)
    s, e = trading_day_window()
    current_stocks = portfolio_df[portfolio_df['shares'] > 0].copy()
    print(f"📊 Processing {len(current_stocks)} current holdings for {today_date}")
    
    for _, stock in current_stocks.iterrows():
        ticker = str(stock["ticker"]).upper()
        shares = float(stock["shares"]) if not pd.isna(stock["shares"]) else 0.0
        cost = float(stock["buy_price"]) if not pd.isna(stock["buy_price"]) else 0.0
        cost_basis = float(stock["cost_basis"]) if not pd.isna(stock["cost_basis"]) else cost * shares
        stop = float(stock["stop_loss"]) if not pd.isna(stock["stop_loss"]) else 0.0

        fetch = download_price_data(ticker, start=s, end=e, auto_adjust=False, progress=False)
        data = fetch.df

        if data.empty:
            print(f"No data for {ticker} (source={fetch.source}).")
            row = {
                "Date": today_date.strftime("%Y-%m-%d"), "Ticker": ticker, "Shares": shares,
                "Buy Price": cost, "Cost Basis": None, "Stop Loss": stop,
                "Current Price": None, "Total Value": None, "PnL": None, "PnL %": None,
                "Action": "NO DATA", "Cash Balance": 0, "Total Equity": None, "Notes": None,
            }
            results.append(row)
            continue

        o = float(data["Open"].iloc[-1]) if "Open" in data else np.nan
        h = float(data["High"].iloc[-1])
        l = float(data["Low"].iloc[-1])
        c = float(data["Close"].iloc[-1])
        if np.isnan(o):
            o = c

        if stop and l <= stop:
            exec_price = round(o if o <= stop else stop, 2)
            action = "SELL - Stop Loss Triggered"
            cash_from_sale = exec_price * shares  # Calculate cash from sale
            cash += cash_from_sale  # Update total cash
            portfolio_df = log_sell(ticker, shares, exec_price, cost, exec_price * shares - cost * shares, portfolio_df)
            row = {
                "Date": today_date.strftime("%Y-%m-%d"), "Ticker": ticker, "Shares": shares,  # Show shares sold (not 0)
                "Buy Price": cost, "Cost Basis": None, "Stop Loss": stop,
                "Current Price": exec_price, "Total Value": None, "PnL": None, "PnL %": None,
                "Action": action, "Cash Balance": None, "Total Equity": None, "Notes": None,
            }
        else:
            price = round(c, 2)
            
            # Determine if this is a BUY (new purchase today) or HOLD (existing position)
            # Check if this ticker was purchased today by looking at trade log
            action = "HOLD"  # Default for existing positions
            trade_log_path = data_dir / "chatgpt_trade_log.csv"
            if trade_log_path.exists():
                trade_log = pd.read_csv(trade_log_path)
                today_buys = trade_log[
                    (trade_log['Date'] == today_date.strftime('%Y-%m-%d')) & 
                    (trade_log['Ticker'] == ticker) & 
                    (trade_log['Shares Bought'].notna())
                ]
                if not today_buys.empty:
                    action = "BUY"  # This stock was purchased today
            
            # Check if this stock had a manual sell and update cash balance accordingly
            cash_balance = manual_sells.get(ticker, 0)
            # Don't calculate totals here - let Excel formulas handle it
            row = {
                "Date": today_date.strftime("%Y-%m-%d"), "Ticker": ticker, "Shares": shares,
                "Buy Price": cost, "Cost Basis": None, "Stop Loss": stop,
                "Current Price": price, "Total Value": None, "PnL": None, "PnL %": None,
                "Action": action, "Cash Balance": round(cash_balance, 2), "Total Equity": None, "Notes": None,
            }

        results.append(row)

    # Add TOTAL row - let Excel SUM formulas calculate the totals
    total_row = {
        "Date": today_date.strftime("%Y-%m-%d"), "Ticker": "TOTAL", "Shares": None, "Buy Price": None,
        "Cost Basis": None, "Stop Loss": None, "Current Price": None,
        "Total Value": None, "PnL": None, "PnL %": None,  # Leave empty for Excel formulas
        "Action": None, "Cash Balance": round(cash, 2),
        "Total Equity": None, "Notes": None,  # Leave empty for Excel formulas
    }
    results.append(total_row)

    # Save results to Excel (preserving existing data and formulas)
    new_results_df = pd.DataFrame(results)
    save_portfolio_results_excel(new_results_df, data_dir)

    return portfolio_df, cash

def save_portfolio_results_excel(results_df: pd.DataFrame, data_dir: Path) -> None:
    """Save portfolio results to Excel file while preserving existing formulas."""
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    excel_path = data_dir / "chatgpt_portfolio_update.xlsx"
    today_date = last_trading_date().date()
    
    if not excel_path.exists():
        # If file doesn't exist, create it
        success = write_portfolio_excel(results_df, excel_path)
        if success:
            logger.info(f"✅ Created new Excel portfolio file: {excel_path}")
        return
    
    try:
        # Load existing workbook to preserve formulas
        wb = load_workbook(excel_path)
        ws = wb['Portfolio']
        
        # Find rows with today's date and remove them
        rows_to_delete = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_date = row[0]
            if isinstance(row_date, str):
                row_date = pd.to_datetime(row_date).date()
            elif hasattr(row_date, 'date'):
                row_date = row_date.date()
            
            if row_date == today_date:  # Date column
                rows_to_delete.append(row_idx)
        
        # Delete rows in reverse order to maintain indices
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)
            logger.info(f"Removed existing data for {today_date} from row {row_idx}")
        
        # Add new data at the end and insert formulas
        start_row = ws.max_row + 1
        for row_idx, row in enumerate(dataframe_to_rows(results_df, index=False, header=False)):
            current_row = start_row + row_idx
            ws.append(row)
            
            # Add formulas for individual stock rows (not TOTAL rows)
            if row[1] != 'TOTAL':  # Ticker column
                # Cost Basis = 0 for SELL, Shares * Buy Price for others
                ws[f'E{current_row}'] = f'=IF(ISNUMBER(SEARCH("SELL",K{current_row})),0,IF(AND(C{current_row}<>"",D{current_row}<>""),C{current_row}*D{current_row},""))'
                # Total Value = Shares * Current Price (C * G)
                ws[f'H{current_row}'] = f'=IF(AND(C{current_row}<>"",G{current_row}<>""),C{current_row}*G{current_row},"")'
                # P&L = (Current Price - Buy Price) * Shares
                ws[f'I{current_row}'] = f'=IF(AND(C{current_row}<>"",D{current_row}<>"",G{current_row}<>""),C{current_row}*(G{current_row}-D{current_row}),"")'
                # P&L % = (Current Price - Buy Price) / Buy Price * 100
                ws[f'J{current_row}'] = f'=IF(AND(D{current_row}<>"",D{current_row}<>0,G{current_row}<>""),((G{current_row}-D{current_row})/D{current_row})*100,"")'
                # Cash Balance = +Total Value for SELL, -Total Value for BUY, 0 for HOLD
                ws[f'L{current_row}'] = f'=IF(ISNUMBER(SEARCH("SELL",K{current_row})),H{current_row},IF(ISNUMBER(SEARCH("BUY",K{current_row})),-H{current_row},0))'
                # Total Equity = IF SELL transaction, use P&L, otherwise Total Value (market value)
                ws[f'M{current_row}'] = f'=IF(ISNUMBER(SEARCH("SELL",K{current_row})),I{current_row},H{current_row})'
            else:
                # TOTAL row formulas - need to find the range of individual stock rows for today
                stock_rows = []
                for r in range(start_row, current_row):
                    if ws[f'B{r}'].value != 'TOTAL':  # Find non-TOTAL rows
                        stock_rows.append(r)
                
                if stock_rows:
                    first_stock_row = min(stock_rows)
                    last_stock_row = max(stock_rows)
                    # Total Cost Basis = SUM of individual cost basis
                    ws[f'E{current_row}'] = f'=SUM(E{first_stock_row}:E{last_stock_row})'
                    # Total Value = SUM of individual stock total values
                    ws[f'H{current_row}'] = f'=SUM(H{first_stock_row}:H{last_stock_row})'
                    # Total P&L = SUM of individual stock P&L
                    ws[f'I{current_row}'] = f'=SUM(I{first_stock_row}:I{last_stock_row})'
                    # Total Cash = SUM of individual stock cash (should mostly be 0 except for sales)
                    ws[f'L{current_row}'] = f'=SUM(L{first_stock_row}:L{last_stock_row})'
                    # Total Equity = SUM of individual stock equity
                    ws[f'M{current_row}'] = f'=SUM(M{first_stock_row}:M{last_stock_row})'
                
                # Apply grey background and bold font to TOTAL rows
                from openpyxl.styles import PatternFill, Font
                grey_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                bold_font = Font(bold=True)
                
                for col_num in range(1, ws.max_column + 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.fill = grey_fill
                    cell.font = bold_font
        
        # Save workbook (preserves formulas)
        wb.save(excel_path)
        logger.info(f"✅ Portfolio results saved to Excel: {excel_path} (formulas preserved)")
        
    except Exception as e:
        logger.error(f"❌ Failed to save portfolio results to Excel: {e}")
        # Fallback to standard method
        existing_df = read_portfolio_excel(excel_path)
        if not existing_df.empty:
            # Convert dates for comparison
            existing_df['Date'] = pd.to_datetime(existing_df['Date'], errors='coerce')
            existing_df = existing_df[existing_df['Date'].dt.date != today_date]
        
        combined_df = pd.concat([existing_df, results_df], ignore_index=True) if not existing_df.empty else results_df
        write_portfolio_excel(combined_df, excel_path)

def daily_results_excel(chatgpt_portfolio: pd.DataFrame, cash: float, data_dir: Path) -> None:
    """Generate comprehensive daily results directly from Excel data."""
    from trading_script import load_benchmarks, last_trading_date, download_price_data, check_weekend
    import numpy as np
    
    portfolio_dict: list[dict] = chatgpt_portfolio.to_dict(orient="records")
    today = check_weekend()
    
    # Price & Volume section
    rows: list[list[str]] = []
    header = ["Ticker", "Close", "% Chg", "Volume"]
    
    end_d = last_trading_date()
    start_d = (end_d - pd.Timedelta(days=4)).normalize()
    
    benchmarks = load_benchmarks()
    benchmark_entries = [{"ticker": t} for t in benchmarks]
    
    for stock in portfolio_dict + benchmark_entries:
        ticker = str(stock["ticker"]).upper()
        try:
            fetch = download_price_data(ticker, start=start_d, end=(end_d + pd.Timedelta(days=1)), progress=False)
            data = fetch.df
            if data.empty or len(data) < 2:
                rows.append([ticker, "—", "—", "—"])
                continue
            
            price = float(data["Close"].iloc[-1])
            last_price = float(data["Close"].iloc[-2])
            volume = float(data["Volume"].iloc[-1])
            
            percent_change = ((price - last_price) / last_price) * 100
            rows.append([ticker, f"{price:,.2f}", f"{percent_change:+.2f}%", f"{int(volume):,}"])
        except Exception as e:
            rows.append([ticker, "—", "—", "—"])
    
    # Read Excel portfolio history for risk calculations
    excel_df = read_portfolio_excel(data_dir / "chatgpt_portfolio_update.xlsx")
    
    # Get TOTAL rows and calculate Total Equity values
    totals = excel_df[excel_df['Ticker'] == 'TOTAL'].copy()
    
    if not totals.empty:
        # Calculate Total Equity for each TOTAL row based on the date's individual stocks
        for idx, total_row in totals.iterrows():
            date = total_row['Date']
            
            # Get all stock rows for this date
            date_stocks = excel_df[(excel_df['Date'] == date) & (excel_df['Ticker'] != 'TOTAL')]
            
            total_equity = 0.0
            for _, stock_row in date_stocks.iterrows():
                # Calculate equity for this stock
                action = str(stock_row.get('Action', '')).upper()
                shares = float(stock_row.get('Shares', 0)) if pd.notna(stock_row.get('Shares')) else 0
                buy_price = float(stock_row.get('Buy Price', 0)) if pd.notna(stock_row.get('Buy Price')) else 0
                current_price = float(stock_row.get('Current Price', 0)) if pd.notna(stock_row.get('Current Price')) else 0
                cash_balance = float(stock_row.get('Cash Balance', 0)) if pd.notna(stock_row.get('Cash Balance')) else 0
                
            if 'SELL' in action:
                # For SELL: equity = realized P&L (negative for loss)
                # shares here represents shares sold, buy_price is original purchase price
                cost_basis = shares * buy_price
                cash_received = shares * current_price  # current_price is the sell price
                realized_pnl = cash_received - cost_basis  # This will be negative for a loss
                stock_equity = realized_pnl  # Total Equity = realized P&L
            else:
                # For others: equity = current value + cash
                total_value = shares * current_price
                stock_equity = total_value + cash_balance
                
                total_equity += stock_equity
            
            # Add any portfolio-level cash (from Cash Balance column of TOTAL row)
            portfolio_cash = float(total_row.get('Cash Balance', 0)) if pd.notna(total_row.get('Cash Balance')) else 0
            total_equity += portfolio_cash
            
            # Update the TOTAL row's Total Equity
            excel_df.loc[idx, 'Total Equity'] = total_equity
        
        # Convert dates and create equity time series
        totals = excel_df[excel_df['Ticker'] == 'TOTAL'].copy()
        totals['Date'] = pd.to_datetime(totals['Date'], format="mixed", errors="coerce")
        totals = totals.dropna(subset=['Date']).sort_values('Date')
        
        if len(totals) > 0:
            final_equity = float(totals.iloc[-1]['Total Equity']) if pd.notna(totals.iloc[-1]['Total Equity']) else 0.0
            equity_series = totals.set_index('Date')['Total Equity'].astype(float).sort_index()
            
            # Calculate risk metrics
            running_max = equity_series.cummax()
            drawdowns = (equity_series / running_max) - 1.0
            max_drawdown = float(drawdowns.min()) if not drawdowns.empty else 0.0
            mdd_date = drawdowns.idxmin() if not drawdowns.empty else pd.NaT
            
            # Daily returns for Sharpe/Sortino
            r = equity_series.pct_change().dropna()
            if len(r) >= 2:
                rf_annual = 0.045
                rf_daily = (1 + rf_annual) ** (1 / 252) - 1
                
                mean_daily = float(r.mean())
                std_daily = float(r.std(ddof=1))
                
                # Risk metrics
                sharpe_annual = ((mean_daily - rf_daily) / std_daily) * np.sqrt(252) if std_daily > 0 else np.nan
                
                downside = (r - rf_daily).clip(upper=0)
                downside_std = float((downside.pow(2).mean()) ** 0.5) if not downside.empty else np.nan
                sortino_annual = ((mean_daily - rf_daily) / downside_std) * np.sqrt(252) if downside_std and downside_std > 0 else np.nan
            else:
                sharpe_annual = np.nan
                sortino_annual = np.nan
        else:
            final_equity = 0.0
            max_drawdown = 0.0
            mdd_date = pd.NaT
            sharpe_annual = np.nan
            sortino_annual = np.nan
    else:
        final_equity = 0.0
        max_drawdown = 0.0
        mdd_date = pd.NaT
        sharpe_annual = np.nan
        sortino_annual = np.nan
    
    # Format output
    def fmt_or_na(x: float, fmt: str) -> str:
        return (fmt.format(x) if not (x is None or (isinstance(x, float) and np.isnan(x))) else "N/A")
    
    # Print comprehensive report
    print("\n" + "=" * 64)
    print(f"Daily Results — {today}")
    print("=" * 64)
    
    print("\n[ Price & Volume ]")
    colw = [10, 12, 9, 15]
    print(f"{header[0]:<{colw[0]}} {header[1]:>{colw[1]}} {header[2]:>{colw[2]}} {header[3]:>{colw[3]}}")
    print("-" * sum(colw) + "-" * 3)
    for rrow in rows:
        print(f"{str(rrow[0]):<{colw[0]}} {str(rrow[1]):>{colw[1]}} {str(rrow[2]):>{colw[2]}} {str(rrow[3]):>{colw[3]}}")
    
    print("\n[ Risk & Return ]")
    mdd_date_str = mdd_date.date() if hasattr(mdd_date, 'date') else str(mdd_date)
    print(f"{'Max Drawdown:':32} {fmt_or_na(max_drawdown, '{:.2%}'):>15}   on {mdd_date_str}")
    print(f"{'Sharpe Ratio (annualized):':32} {fmt_or_na(sharpe_annual, '{:.4f}'):>15}")
    print(f"{'Sortino Ratio (annualized):':32} {fmt_or_na(sortino_annual, '{:.4f}'):>15}")
    
    print("\n[ CAPM vs Benchmarks ]")
    print("Beta/Alpha: insufficient overlapping data.")  # Simplified for now
    
    print("\n[ Snapshot ]")
    print(f"{'Latest ChatGPT Equity:':32} ${final_equity:>14,.2f}")
    print(f"{'Cash Balance:':32} ${cash:>14,.2f}")
    
    print("\n[ Holdings ]")
    # Create a detailed holdings table
    holdings_data = []
    for _, stock in chatgpt_portfolio.iterrows():
        ticker = stock['ticker']
        shares = stock['shares']
        buy_price = stock['buy_price']
        stop_loss = stock['stop_loss']
        
        # Get current price from the price data we already fetched
        current_price = 0.0
        for row in rows:
            if row[0] == ticker.upper():
                try:
                    current_price = float(str(row[1]).replace(',', ''))
                except:
                    current_price = 0.0
                break
        
        if current_price > 0:
            pnl_pct = ((current_price - buy_price) / buy_price) * 100
        else:
            pnl_pct = 0.0
        
        holdings_data.append({
            'ticker': ticker,
            'shares': shares,
            'buy_price': buy_price,
            'cost_basis': buy_price * shares,
            'stop_loss': stop_loss,
            'PnL %': f"{pnl_pct:.2f}%",
            'Notes': None
        })
    
    holdings_df = pd.DataFrame(holdings_data)
    print(holdings_df.to_string(index=True))
    
def main_excel(data_dir: Path = None):
    """Main function for Excel-based trading script.
    
    Args:
        data_dir: Path to data directory
    """
    
    # Set up data directory
    if data_dir:
        set_data_dir(data_dir)
        data_directory = data_dir
    else:
        data_directory = Path("Start Your Own")
    
    logger.info(f"Using data directory: {data_directory}")
    
    # Load portfolio state
    try:
        chatgpt_portfolio, cash = load_latest_portfolio_state_excel(data_directory)
    except FileNotFoundError as e:
        print(str(e))
        return
    
    # Always run interactive mode (like original script)
    updated_portfolio, updated_cash = process_portfolio_excel(chatgpt_portfolio, cash, data_directory, interactive=True)
    daily_results_excel(updated_portfolio, updated_cash, data_directory)

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Excel-compatible Portfolio Trading Script")
    parser.add_argument("--data-dir", type=str, help="Data directory path")
    parser.add_argument("--log-level", type=str, default="INFO", choices=["DEBUG", "INFO", "WARNING", "ERROR"])
    
    args = parser.parse_args()
    
    # Set log level
    logging.getLogger().setLevel(getattr(logging, args.log_level))
    
    # Log initial state
    _log_initial_state()
    
    logger.info(f"Script started with arguments: {vars(args)}")
    
    # Run main function (always interactive like original script)
    main_excel(
        data_dir=Path(args.data_dir) if args.data_dir else None
    )
