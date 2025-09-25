"""
Excel utilities for portfolio management.
These functions handle reading/writing portfolio data to/from Excel files.
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

logger = logging.getLogger(__name__)

def read_portfolio_excel(excel_path: Path) -> pd.DataFrame:
    """Read portfolio data from Excel file."""
    if not excel_path.exists():
        logger.warning(f"Excel file not found: {excel_path}")
        return pd.DataFrame()
    
    try:
        logger.info(f"Reading Excel file: {excel_path}")
        df = pd.read_excel(excel_path, sheet_name='Portfolio')
        logger.info(f"Successfully read Excel file: {excel_path}")
        return df
    except Exception as e:
        logger.error(f"Error reading Excel file {excel_path}: {e}")
        return pd.DataFrame()

def write_portfolio_excel(df: pd.DataFrame, excel_path: Path) -> bool:
    """Write portfolio data to Excel file with formatting."""
    try:
        logger.info(f"Writing Excel file: {excel_path}")
        
        # Create or load workbook
        if excel_path.exists():
            wb = load_workbook(excel_path)
            if 'Portfolio' in wb.sheetnames:
                ws = wb['Portfolio']
                # Clear existing data
                ws.delete_rows(1, ws.max_row)
            else:
                ws = wb.create_sheet('Portfolio')
                if 'Sheet' in wb.sheetnames:  # Remove default sheet
                    wb.remove(wb['Sheet'])
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Portfolio'
        
        # Add data
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format header row
        if ws.max_row > 0:
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:  # Header row
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Format TOTAL rows
            for row in ws.iter_rows(min_row=2):  # Skip header
                if row[1].value == "TOTAL":  # Ticker column
                    for cell in row:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save
        wb.save(excel_path)
        logger.info(f"Successfully wrote Excel file: {excel_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error writing Excel file {excel_path}: {e}")
        return False

def append_to_portfolio_excel(new_rows: pd.DataFrame, excel_path: Path) -> bool:
    """Append new rows to existing Excel portfolio file."""
    try:
        # Read existing data
        existing_df = read_portfolio_excel(excel_path)
        
        # Combine with new data
        if existing_df.empty:
            combined_df = new_rows
        else:
            combined_df = pd.concat([existing_df, new_rows], ignore_index=True)
        
        # Write back
        return write_portfolio_excel(combined_df, excel_path)
        
    except Exception as e:
        logger.error(f"Error appending to Excel file {excel_path}: {e}")
        return False